#!/usr/bin/env python3
"""HANABI 原価管理表 (Box xlsx) をパース → data/cost_ratios.json 出力。

入力: Box の固定パス (複数 FY を全部スキャン → マージ)
  - HANABI_原価計算表_ver0.9.xlsx (FY25 = 2025/05〜2026/04, 過去データ)
  - HABABI_FY26予実管理表.xlsx (FY26+ = 2026/05〜, 現行更新中)
  注: FY26ファイル名の "HABABI" タイポは原本のまま

出力: data/cost_ratios.json
  {
    "targets": {fy: {sid: ratio}},  # FY別の規定値
    "monthly": {YYYYMM: {sid: {sales, cost, ratio}}},  # 全FY 月別データを統合
    "fy_average": {fy: {sid: ratio}},
    "notes": {...},
    "source": {parsed_at, files: [...]}
  }

セクションキー (sid):
  - tsunashima (Hanabi 綱島店、 ヘア)
  - miyakojima_total (ELLE宮古島 店舗合計)
  - miyakojima_hair
  - miyakojima_eye
  - miyakojima_nail
"""
from __future__ import annotations

import json
import re
import sys
from datetime import datetime
from pathlib import Path
from zoneinfo import ZoneInfo

from openpyxl import load_workbook

ROOT = Path(__file__).resolve().parent.parent
DATA_DIR = ROOT / "data"
OUTPUT = DATA_DIR / "cost_ratios.json"
JST = ZoneInfo("Asia/Tokyo")

# Box ファイル候補 (複数 FY を全部パースして統合)
# 🛡 2026-07-03: ファイルが 010_運営本部/ 直下 → 000_本部/ 配下に移動 + ver サフィックス付きに
#   リネームされて月初取込が silent skip していた。 glob で「名前の揺れ (ver0.5等)・置き場所」 に
#   耐える方式に変更。 同名パターン複数ヒット時は名前降順 (=最新ver) を採用。
_BOX_BASE = Path.home() / "Library/CloudStorage/Box-Box/050_HANABI/010_運営本部"
def _find_box_files():
    patterns = ["HANABI_原価計算表*.xlsx", "HABABI_FY*予実管理表*.xlsx", "HANABI_FY*予実管理表*.xlsx"]
    found = []
    for sub in [_BOX_BASE, _BOX_BASE / "000_本部"]:
        for pat in patterns:
            for p in sorted(sub.glob(pat), reverse=True):
                # old/ フォルダは対象外、 同種 (パターン×FY部分) は最新verのみ
                key = (pat, p.name.split("_ver")[0])
                if key not in [(k, n) for k, n, _ in found]:
                    found.append((pat, p.name.split("_ver")[0], p))
    return [p for _, _, p in found]
BOX_PATHS = _find_box_files()


def safe_num(v):
    if v is None:
        return None
    if isinstance(v, (int, float)):
        if isinstance(v, float) and v != v:
            return None
        return v
    return None


def yyyymm_iter(start: str, end: str):
    y, m = int(start[:4]), int(start[4:6])
    ey, em = int(end[:4]), int(end[4:6])
    while (y, m) <= (ey, em):
        yield f"{y:04d}{m:02d}"
        m += 1
        if m == 13:
            m = 1
            y += 1


def find_cost_sheet(wb) -> tuple[str, str] | tuple[None, None]:
    """原価管理表 / 原価計算表 / 「YYYY年度」 シートを見つけて FY を抽出"""
    candidates = []
    for sn in wb.sheetnames:
        # FY26 系: シート名に 'FY##' と '原価' を含む
        if "原価" in sn:
            m = re.search(r"FY(\d{2})", sn)
            if m:
                return sn, f"FY{m.group(1)}"
            # 「2025年度」 形式
            m = re.search(r"(\d{4})年度", sn)
            if m:
                yy = int(m.group(1)) % 100
                return sn, f"FY{yy:02d}"
            candidates.append((sn, None))
        # 「2025年度」 単独
        m = re.search(r"(\d{4})年度", sn)
        if m:
            yy = int(m.group(1)) % 100
            candidates.append((sn, f"FY{yy:02d}"))
    if candidates:
        return candidates[0]
    return (wb.sheetnames[0], None) if wb.sheetnames else (None, None)


def parse_period(ws, fy_hint: str | None) -> tuple[str, str]:
    """対象期間 文字列 or fy_hint から (YYYYMM_start, YYYYMM_end) を返す"""
    for r in range(1, 6):
        for c in range(1, 5):
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str):
                continue
            m = re.search(r"(\d{4})/(\d{1,2})\s*[~〜~]\s*(\d{4})/(\d{1,2})", v)
            if m:
                y1, mo1, y2, mo2 = m.groups()
                return f"{int(y1):04d}{int(mo1):02d}", f"{int(y2):04d}{int(mo2):02d}"
    # fallback: fy_hint から
    if fy_hint and fy_hint.startswith("FY"):
        try:
            yy = int(fy_hint[2:])
            return f"20{yy:02d}05", f"20{yy+1:02d}04"
        except ValueError:
            pass
    return None, None


def find_block_starts(ws, header_row: int | None) -> dict[str, int]:
    """各セクションの「技術売上」行番号を見つける。

    FY25/FY26 共通レイアウト: ヘッダー行 (col 5 = datetime) からの相対オフセット:
      tsunashima         : header_row + 1   (綱島 技術売上)
      miyakojima_total   : header_row + 6   (ELLE宮古島 店舗合計 技術売上)
      miyakojima_hair    : header_row + 10
      miyakojima_eye     : header_row + 14
      miyakojima_nail    : header_row + 18
    """
    if not header_row:
        return {}
    starts = {
        "tsunashima":       header_row + 1,
        "miyakojima_total": header_row + 6,
        "miyakojima_hair":  header_row + 10,
        "miyakojima_eye":   header_row + 14,
        "miyakojima_nail":  header_row + 18,
    }
    # 検証: 各 start_row の col 4 (D) が「技術売上」 を含むか
    valid = {}
    for sid, r in starts.items():
        if r > ws.max_row:
            continue
        c4 = str(ws.cell(row=r, column=4).value or "")
        if "技術売上" in c4:
            valid[sid] = r
        else:
            print(f"  ⚠ {sid} 期待位置 R{r} に「技術売上」 ない (実際: {c4[:30]}) — skip", file=sys.stderr)
    return valid


def find_header_row(ws) -> int | None:
    """月ヘッダー行 (col 5 に datetime 値) を探す"""
    for r in range(1, 15):
        v = ws.cell(row=r, column=5).value
        if isinstance(v, datetime):
            return r
    return None


def parse_targets_from_cells(ws) -> dict:
    """FY26形式: R4-R7 col D に専用セル"""
    t = {
        "tsunashima":      safe_num(ws.cell(row=4, column=4).value),
        "miyakojima_hair": safe_num(ws.cell(row=5, column=4).value),
        "miyakojima_eye":  safe_num(ws.cell(row=6, column=4).value),
        "miyakojima_nail": safe_num(ws.cell(row=7, column=4).value),
    }
    return t if all(v is not None for v in t.values()) else None


def parse_targets_from_text(ws) -> dict:
    """FY25形式: 上部の説明テキストから 正規表現で抽出"""
    targets = {}
    for r in range(1, 10):
        for c in range(1, 6):
            v = ws.cell(row=r, column=c).value
            if not v or not isinstance(v, str):
                continue
            text = v
            # "綱島店：8.5%"
            m = re.search(r"綱島店?[：:]?\s*([\d.]+)\s*%", text)
            if m:
                targets["tsunashima"] = float(m.group(1)) / 100
            # "宮古島店：＜ヘア＞8.5% ＜アイ＞5.0% ＜ネイル＞6%"
            if "宮古島" in text:
                m = re.search(r"ヘア[＞>]?\s*([\d.]+)\s*%", text)
                if m:
                    targets["miyakojima_hair"] = float(m.group(1)) / 100
                m = re.search(r"アイ[＞>]?\s*([\d.]+)\s*%", text)
                if m:
                    targets["miyakojima_eye"] = float(m.group(1)) / 100
                m = re.search(r"ネイル[＞>]?\s*([\d.]+)\s*%", text)
                if m:
                    targets["miyakojima_nail"] = float(m.group(1)) / 100
    return targets


def parse_targets(ws) -> dict:
    cells_t = parse_targets_from_cells(ws)
    if cells_t:
        return cells_t
    return parse_targets_from_text(ws) or {}


def parse_one_file(xlsx_path: Path) -> dict | None:
    """1ファイル パース → {fy, period, targets, monthly{ym:{sid:...}}, fy_average}"""
    try:
        wb = load_workbook(xlsx_path, data_only=True)
    except Exception as e:
        print(f"  ✗ load failed: {e}", file=sys.stderr)
        return None
    sheet_name, fy = find_cost_sheet(wb)
    if not sheet_name:
        print(f"  ✗ 原価シート見つからず", file=sys.stderr)
        return None
    ws = wb[sheet_name]
    print(f"  sheet: {sheet_name} (推定 FY: {fy}), {ws.max_row}r x {ws.max_column}c")

    start_ym, end_ym = parse_period(ws, fy)
    if not start_ym:
        print(f"  ✗ 対象期間 取得失敗", file=sys.stderr)
        return None
    months = list(yyyymm_iter(start_ym, end_ym))
    n_months = len(months)
    print(f"  期間: {start_ym} - {end_ym} ({n_months}ヶ月)")

    targets = parse_targets(ws)
    print(f"  targets: {targets}")

    header_r = find_header_row(ws)
    print(f"  header_row: R{header_r}")
    block_starts = find_block_starts(ws, header_r)
    print(f"  block_starts: {block_starts}")

    # データ開始列: ヘッダー行から実際に datetime 値が入ってる先頭列を探す
    data_start_col = 5
    if header_r:
        for c in range(5, ws.max_column + 1):
            if isinstance(ws.cell(row=header_r, column=c).value, datetime):
                data_start_col = c
                break

    monthly_out = {ym: {} for ym in months}
    fy_average = {}
    for sid, start_row in block_starts.items():
        # 4行ブロック: 技術売上 / 原価 / 売上原価 / 割合
        sales_row, cost_row, _, ratio_row = start_row, start_row + 1, start_row + 2, start_row + 3
        for i, ym in enumerate(months):
            col = data_start_col + i
            if col > ws.max_column:
                break
            sales = safe_num(ws.cell(row=sales_row, column=col).value)
            cost = safe_num(ws.cell(row=cost_row, column=col).value)
            ratio = safe_num(ws.cell(row=ratio_row, column=col).value)
            if ratio is None and sales and cost is not None:
                ratio = cost / sales if sales > 0 else None
            # 全部 None or 0 ならスキップ (空月)
            if (sales is None or sales == 0) and (cost is None or cost == 0):
                continue
            monthly_out[ym][sid] = {
                "sales": int(sales) if sales is not None else None,
                "cost": int(cost) if cost is not None else None,
                "ratio": round(ratio, 6) if ratio is not None else None,
            }
        # FY平均 (data_start_col + n_months 列に「平均」)
        avg_col = data_start_col + n_months
        if avg_col <= ws.max_column:
            avg_ratio = safe_num(ws.cell(row=ratio_row, column=avg_col).value)
            if avg_ratio is not None:
                fy_average[sid] = round(avg_ratio, 6)

    # 空月を除外
    monthly_out = {ym: d for ym, d in monthly_out.items() if d}
    return {
        "fy": fy,
        "period": {"start": start_ym, "end": end_ym},
        "targets": {k: round(v, 6) if v is not None else None for k, v in targets.items()},
        "monthly": monthly_out,
        "fy_average": fy_average,
    }


def main() -> int:
    DATA_DIR.mkdir(exist_ok=True)
    all_monthly = {}
    all_targets = {}   # {fy: {sid: ratio}}
    all_avg = {}       # {fy: {sid: ratio}}
    sources = []

    found_any = False
    for path in BOX_PATHS:
        if not path.exists():
            print(f"[skip] {path.name} 不存在")
            continue
        found_any = True
        print(f"[parse] {path.name}")
        result = parse_one_file(path)
        if not result:
            continue
        fy = result["fy"] or "?"
        all_targets[fy] = result["targets"]
        all_avg[fy] = result["fy_average"]
        # マージ (新しい FY のデータが優先で上書き)
        all_monthly.update(result["monthly"])
        sources.append({"file": str(path), "fy": fy, "months": len(result["monthly"])})

    if not found_any:
        print(f"[parse_cost_ratios] ⚠️ Box xlsx 全て不存在 — Box Drive 起動 or パス確認", file=sys.stderr)
        # 既存 JSON があれば 維持 (上書きで履歴消えるのを防ぐ)
        if OUTPUT.exists():
            print(f"  → 既存 {OUTPUT.name} 維持")
            return 0
        # 初回で何も無い場合のみ 空 schema を書く
        out = {"_comment": "Box xlsx 未取得", "targets": {}, "monthly": {}, "fy_average": {}, "sources": []}
        OUTPUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
        return 0

    # 🛡 Box 接続できたが パース結果が空 (構造変化 or 一時障害) なら 既存維持
    if not all_monthly and OUTPUT.exists():
        try:
            existing = json.loads(OUTPUT.read_text(encoding="utf-8"))
            if existing.get("monthly"):
                print(f"[parse_cost_ratios] ⚠️ パース結果が空 — 既存データ ({len(existing['monthly'])}ヶ月) を維持して上書きしない", file=sys.stderr)
                return 0
        except Exception:
            pass

    out = {
        "_comment": "HANABI 原価管理表 (Box xlsx) から自動生成。 scripts/parse_cost_ratios.py。 手動編集禁止 (上書き)。",
        "targets": all_targets,
        "monthly": {ym: all_monthly[ym] for ym in sorted(all_monthly.keys())},
        "fy_average": all_avg,
        "notes": {
            "hair": "※HAIRはエクステを抜いた原価率を転記 (在庫管理表ベース)"
        },
        "sources": sources,
        "parsed_at": datetime.now(JST).isoformat(),
    }
    OUTPUT.write_text(json.dumps(out, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"\n[parse_cost_ratios] ✓ wrote {OUTPUT} ({OUTPUT.stat().st_size} bytes)")
    print(f"  月数: {len(out['monthly'])}, FY: {list(all_targets.keys())}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
